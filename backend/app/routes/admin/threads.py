# backend/app/routes/admin/threads.py
"""Admin thread routes with security enhancements.

Provides CRUD operations for threads, including input validation,
XSS sanitization, and secure file upload handling.
"""

from flask import (
    render_template,
    request,
    redirect,
    url_for,
    flash,
    current_app,
    session,
)
import bleach

from . import admin_bp
from ...models import Thread, Project, Message, User
from ... import logic, tasks
from ...extensions import db, limiter
from .security import log_audit
from .forms import AddThreadForm


@admin_bp.route('/threads/view/<thread_id>')
def view_thread(thread_id):
    if not session.get('user_id'):
        return redirect(url_for('auth.login'))

    group_id = request.args.get('group_id')
    project = None
    if group_id:
        project = Project.query.get(group_id)
    else:
        t = Thread.query.filter_by(thread_id=thread_id).first()
        if t:
            project = t.project

    if not project:
        flash('Project not found for this thread', 'error')
        return redirect(url_for('admin.index'))

    is_owner = any(o.id == session.get('user_id') for o in project.owners)
    if session.get('role') != 'admin' and not is_owner:
        flash('Permission Denied', 'error')
        return redirect(url_for('admin.index'))

    thread_obj = Thread.query.filter_by(thread_id=thread_id).first()
    is_syncing = False
    result = None

    if thread_obj:
        try:
            tasks.refresh_specific_threads.schedule(
                args=(project.id, [thread_id]), delay=0)
            if not request.args.get('nomsg'):
                flash('正在後台更新數據，頁面顯示為快取資料。', 'info')
        except Exception as e:
            current_app.logger.warning(
                f"Failed to trigger async task: {e}")
        result = logic.process_thread_from_db(
            thread_obj, target_name="", start_date=None, end_date=None)
        # Extract remark for template (it's in result['data']['remark'])
        if not result.get('remark') and result.get('data'):
            result['remark'] = result['data'].get('remark', thread_obj.remark or '')
    else:
        try:
            tasks.refresh_specific_threads.schedule(
                args=(project.id, [thread_id]), delay=0)
            is_syncing = True
            result = {
                'data': {'thread_id': thread_id},
                'remark': '',
                'messages': [],
            }
        except Exception as e:
            flash(f'無法啟動同步任務: {e}', 'error')
            return redirect(url_for('admin.index', group_id=project.id))

    return render_template(
        'admin_thread_view.html',
        result=result,
        project=project,
        active_group={'group_id': project.id, 'name': project.name},
        is_syncing=is_syncing,
    )


@admin_bp.route('/threads/add_one', methods=['POST'])
def add_one_thread():
    if not session.get('user_id'):
        return redirect(url_for('auth.login'))

    form = AddThreadForm(request.form)
    if not form.validate():
        error_msg = '; '.join(
            [f"{field}: {msg[0]}" for field, msg in form.errors.items()]
        )
        current_app.logger.warning('Add thread validation error: %s', error_msg)
        flash(f'表單錯誤: {error_msg}', 'error')
        return redirect(url_for('admin.index', group_id=form.group_id.data))

    thread_id = form.thread_id.data
    remark_raw = form.remark.data
    remark = (
        bleach.clean(remark_raw, tags=[], attributes={}, strip=True)
        if remark_raw
        else None
    )
    project = Project.query.get(form.group_id.data)
    if not project:
        return redirect(url_for('admin.index'))

    is_owner = any(o.id == session.get('user_id') for o in project.owners)
    if session.get('role') != 'admin' and not is_owner:
        flash('權限不足', 'error')
        return redirect(url_for('admin.index', group_id=project.id))

    exists = any(t.thread_id == thread_id for t in project.threads)
    if exists:
        flash('Thread ID already exists in this project', 'error')
    else:
        new_t = Thread(
            thread_id=thread_id, project_id=project.id, remark=remark
        )
        db.session.add(new_t)
        project.version += 1
        db.session.commit()
        log_audit('Add Thread', f"User added {thread_id} to {project.name}")
        flash('Thread added with remark' if remark else 'Thread added', 'success')

    return redirect(url_for('admin.index', group_id=project.id))


@admin_bp.route('/threads/delete_multi', methods=['POST'])
def delete_multi():
    if not session.get('user_id'):
        return redirect(url_for('auth.login'))

    group_id = request.form.get('group_id')
    project = Project.query.get(group_id)
    if not project:
        return redirect(url_for('admin.index'))

    is_owner = any(o.id == session.get('user_id') for o in project.owners)
    if session.get('role') != 'admin' and not is_owner:
        flash('權限不足', 'error')
        return redirect(url_for('admin.index', group_id=group_id))

    select_all_pages = request.form.get('select_all_pages') == 'true'
    count = 0

    try:
        if select_all_pages:
            search_q = request.form.get('search_q', '').strip()
            status_filter = request.form.get('status_filter', '').strip()
            
            base_query = db.session.query(Thread.id).filter_by(project_id=project.id)
            if search_q:
                base_query = base_query.filter(
                    (Thread.thread_id.contains(search_q)) |
                    (Thread.remark.contains(search_q))
                )
            if status_filter and status_filter != 'all':
                if status_filter == 'active':
                    base_query = base_query.filter(Thread.refresh_priority.notin_(['low', 'frozen']))
                else:
                    base_query = base_query.filter_by(refresh_priority=status_filter)
                
            # Fix SAWarning: Coercing Subquery object into a select() for use in IN()
            # We select the ID column explicitly from the subquery logic
            # threads_subquery is already a subquery object, but we need to ensure we select the column from it
            # actually, standard modern SA usage: select(Thread.id).where(...)
            
            stmt = db.session.query(Thread.id).filter_by(project_id=project.id)
            if search_q:
                stmt = stmt.filter(
                    (Thread.thread_id.contains(search_q)) |
                    (Thread.remark.contains(search_q))
                )
            if status_filter and status_filter != 'all':
                if status_filter == 'active':
                    stmt = stmt.filter(Thread.refresh_priority.notin_(['low', 'frozen']))
                else:
                    stmt = stmt.filter_by(refresh_priority=status_filter)
            
            # Use subquery? or just use the query object directly if supported?
            # Creating a subquery() and accessing columns is safer
            subq = stmt.subquery()
            
            Message.query.filter(
                Message.thread_id.in_(db.select(subq.c.id))
            ).delete(synchronize_session=False)
            
            count = Thread.query.filter(
                Thread.id.in_(db.select(subq.c.id))
            ).delete(synchronize_session=False)
        else:
            thread_ids = request.form.getlist('selected_ids')
            if not thread_ids:
                single_id = request.form.get('thread_id')
                if single_id:
                    thread_ids = [single_id]
            if not thread_ids:
                flash('No threads selected', 'warning')
                return redirect(url_for('admin.index', group_id=group_id))
            
            # Safety limit to prevent accidental mass deletion
            if len(thread_ids) > 1000:
                flash(f'單次最多刪除1000條記錄，當前選擇了{len(thread_ids)}條', 'error')
                return redirect(url_for('admin.index', group_id=group_id))
            for tid in thread_ids:
                t = Thread.query.filter_by(
                    thread_id=tid, project_id=project.id
                ).first()
                if t:
                    db.session.delete(t)
                    count += 1

        if count > 0:
            project.version += 1
            log_audit('Delete Threads', f"Deleted {count} threads from {project.name}")
        
        db.session.commit()
        flash(f'{count} threads deleted', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Delete Error: {e}")
        flash('刪除失敗，請稍後再試', 'error')

    return redirect(url_for('admin.index', group_id=group_id))


@admin_bp.route('/threads/update_remark', methods=['POST'])
def update_thread_remark():
    if not session.get('user_id'):
        return {'error': 'Unauthorized'}, 401

    data = request.json
    thread_id = data.get('thread_id')
    group_id = data.get('group_id')
    new_remark_raw = data.get('remark', '').strip()
    sanitized_remark = (
        bleach.clean(new_remark_raw, tags=[], attributes={}, strip=True)
        if new_remark_raw
        else None
    )
    if sanitized_remark and len(sanitized_remark) > 200:
        return {'error': 'Remark exceeds maximum length of 200 characters'}, 400

    project = Project.query.get(group_id)
    if not project:
        return {'error': 'Project not found'}, 404

    is_owner = any(o.id == session.get('user_id') for o in project.owners)
    if session.get('role') != 'admin' and not is_owner:
        return {'error': 'Permission denied'}, 403

    thread = Thread.query.filter_by(
        thread_id=thread_id, project_id=project.id
    ).first()
    if thread:
        thread.remark = sanitized_remark
        db.session.commit()
        log_audit('Update Remark', f"Updated remark for {thread_id} in {project.name}")
        return {'success': True}

    return {'error': 'Thread not found'}, 404


@admin_bp.route('/threads/refresh', methods=['POST'])
@limiter.limit("10 per hour")
def refresh_threads_cache():
    if not session.get('user_id'):
        return redirect(url_for('auth.login'))

    group_id = request.form.get('group_id')
    project = Project.query.get(group_id)
    if not project:
        return redirect(url_for('admin.index'))

    is_owner = any(o.id == session.get('user_id') for o in project.owners)
    if session.get('role') != 'admin' and not is_owner:
        flash('Permission Denied', 'error')
        return redirect(url_for('admin.index', group_id=group_id))

    select_all_pages = request.form.get('select_all_pages') == 'true'
    search_q = request.form.get('search_q', '').strip()
    status_filter = request.form.get('status_filter', '').strip()
    thread_ids = []

    if select_all_pages:
        query = Thread.query.with_entities(Thread.thread_id).filter_by(project_id=project.id)
        if search_q:
            query = query.filter(
                (Thread.thread_id.contains(search_q)) |
                (Thread.remark.contains(search_q))
            )
            
        if status_filter and status_filter != 'all':
            if status_filter == 'active':
                query = query.filter(Thread.refresh_priority.notin_(['low', 'frozen']))
            else:
                query = query.filter_by(refresh_priority=status_filter)
            
        thread_ids = [t.thread_id for t in query.all()]
    else:
        thread_ids = request.form.getlist('selected_ids')
        if not thread_ids:
            single = request.form.get('thread_id')
            if single:
                thread_ids = [single]

    if not thread_ids:
        flash('未選擇任何 Thread', 'warning')
        return redirect(url_for('admin.index', group_id=group_id))

    tasks.refresh_specific_threads(
        group_id, thread_ids, group_name=project.name
    )
    log_audit('Refresh Threads', f"Triggered refresh for {len(thread_ids)} threads in {project.name}")
    flash(f'已排程更新 {len(thread_ids)} 筆資料的快取', 'success')
    return redirect(url_for('admin.index', group_id=group_id))


@admin_bp.route('/threads/export', methods=['POST'])
def export_excel():
    if not session.get('user_id'):
        return redirect(url_for('auth.login'))

    group_id = request.form.get('group_id')
    project = Project.query.get(group_id)
    if not project:
        return redirect(url_for('admin.index'))

    is_owner = any(o.id == session.get('user_id') for o in project.owners)
    if session.get('role') != 'admin' and not is_owner:
        flash('Permission Denied', 'error')
        return redirect(url_for('admin.index'))

    from ...services import excel_service
    try:
        select_all_pages = request.form.get('select_all_pages') == 'true'
        search_q = request.form.get('search_q', '').strip()
        filtered_ids = None
        if not select_all_pages:
            ids = request.form.getlist('selected_ids')
            if not ids and request.form.get('thread_id'):
                ids = [request.form.get('thread_id')]
            if ids:
                filtered_ids = ids
            status_filter=request.form.get('status_filter') if select_all_pages else None,
        )
        # Note: excel_service returns a Response object, so we can't easily log *after* it without wrapping.
        # But we can log *before* returning.
        count_str = "All" if select_all_pages else str(len(filtered_ids) if filtered_ids else 0)
        log_audit('Export Excel', f"Exported threads from {project.name} (Count: {count_str})")
        
        return excel_service.generate_excel_export(
            project.id,
            project.name,
            filtered_ids=filtered_ids,
            search_q=search_q if select_all_pages else None,
            status_filter=request.form.get('status_filter') if select_all_pages else None,
        )
    except Exception as e:
        flash(f'Export Failed: {e}', 'error')
        return redirect(url_for('admin.index', group_id=group_id))


@admin_bp.route('/threads/upload', methods=['POST'])
def upload_file():
    if not session.get('user_id'):
        return redirect(url_for('auth.login'))

    group_id = request.form.get('group_id')
    file = request.files.get('file')
    if not file or not file.filename.lower().endswith('.xlsx'):
        flash('請上傳 Excel (.xlsx) 檔案', 'error')
        return redirect(url_for('admin.index', group_id=group_id))

    if file.mimetype != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        flash('檔案類型不正確，請上傳有效的 Excel 檔案', 'error')
        return redirect(url_for('admin.index', group_id=group_id))

    file.seek(0, 2)
    max_size = current_app.config.get('MAX_CONTENT_LENGTH', 2 * 1024 * 1024)
    if file.tell() > max_size:
        flash(
            f'檔案過大，請限制在 {max_size // (1024 * 1024)} MB 以內',
            'error',
        )
        return redirect(url_for('admin.index', group_id=group_id))
    file.seek(0)

    project = Project.query.get(group_id)
    if not project:
        return redirect(url_for('admin.index'))

    client_version = request.form.get('version', type=int)
    current_version = project.version if project.version else 1
    if client_version is not None and client_version != current_version:
        flash('資料已被其他人修改，請重新整理頁面後再試', 'error')
        return redirect(url_for('admin.index', group_id=group_id))

    is_owner = any(o.id == session.get('user_id') for o in project.owners)
    if session.get('role') != 'admin' and not is_owner:
        return redirect(url_for('admin.index'))

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True, data_only=True)
        if 'Threads' not in wb.sheetnames:
            raise ValueError('缺少必需的工作表 "Threads"')
    except Exception as e:
        current_app.logger.error(
            'Excel upload validation failed: %s', e, exc_info=True
        )
        flash('檔案內容無法解析或格式不正確', 'error')
        return redirect(url_for('admin.index', group_id=group_id))

    from ...services import excel_service
    file.seek(0)
    thread_data_map, error = excel_service.parse_excel_for_import(file)
    if error:
        flash(error, 'error')
        return redirect(url_for('admin.index', group_id=group_id))

    action = request.form.get('action', 'add')
    try:
        stats = excel_service.process_import_data(project.id, thread_data_map, action)
        if 'error' in stats:
            flash(f"處理失敗: {stats['error']}", 'error')
        else:
            if action == 'delete':
                count = stats['deleted']
                if count > 0:
                    flash(f'成功刪除 {count} 筆 Thread', 'success')
                    log_audit(
                        'Batch Delete Excel', f"{count} threads from {project.name}"
                    )
                else:
                    flash('沒有刪除任何 Thread', 'warning')
            else:
                added = stats['added']
                updated = stats['updated']
                if added > 0 or updated > 0:
                    flash(
                        f'處理完成: 新增 {added} 筆, 更新 {updated} 筆備註',
                        'success',
                    )
                    log_audit(
                        'Import Excel',
                        f"{added} added, {updated} updated in {project.name}",
                    )
                else:
                    flash('沒有處理任何資料', 'warning')
    except Exception as e:
        current_app.logger.error(
            'Upload processing error: %s', e, exc_info=True
        )
        flash(f'檔案處理失敗: {str(e)}', 'error')

    return redirect(url_for('admin.index', group_id=group_id))


@admin_bp.route('/threads/bookmark/toggle', methods=['POST'])
def toggle_bookmark():
    if not session.get('user_id'):
        return {'error': 'Unauthorized'}, 401
    
    data = request.json
    thread_id = data.get('thread_id')
    
    user = User.query.get(session['user_id'])
    thread = Thread.query.filter_by(thread_id=thread_id).first()
    
    if not thread:
        return {'error': 'Thread not found'}, 404
    
    is_bookmarked = user.bookmarked_threads.filter_by(id=thread.id).first() is not None
    
    if is_bookmarked:
        user.bookmarked_threads.remove(thread)
        action = 'removed'
    else:
        user.bookmarked_threads.append(thread)
        action = 'added'
        
    db.session.commit()
    return {'success': True, 'action': action}

@admin_bp.route('/threads/export-pdf', methods=['POST'])
@limiter.limit("10 per hour")
def export_pdf():
    """Batch PDF export with hybrid sync/async approach"""
    if not session.get('user_id'):
        return redirect(url_for('auth.login'))

    group_id = request.form.get('group_id')
    project = Project.query.get(group_id)
    if not project:
        return redirect(url_for('admin.index'))

    is_owner = any(o.id == session.get('user_id') for o in project.owners)
    if session.get('role') != 'admin' and not is_owner:
        flash('Permission Denied', 'error')
        return redirect(url_for('admin.index'))

    # Parse selection mode
    select_all_pages = request.form.get('select_all_pages') == 'true'
    thread_ids = []

    if select_all_pages:
        search_q = request.form.get('search_q', '').strip()
        status_filter = request.form.get('status_filter', '').strip()
        
        query = Thread.query.with_entities(Thread.thread_id).filter_by(project_id=project.id)
        if search_q:
            query = query.filter(
                (Thread.thread_id.contains(search_q)) |
                (Thread.remark.contains(search_q))
            )
        if status_filter and status_filter != 'all':
            if status_filter == 'active':
                query = query.filter(Thread.refresh_priority.notin_(['low', 'frozen']))
            else:
                query = query.filter_by(refresh_priority=status_filter)
        
        thread_ids = [t.thread_id for t in query.all()]
    else:
        thread_ids = request.form.getlist('selected_ids')
        if not thread_ids:
            single = request.form.get('thread_id')
            if single:
                thread_ids = [single]

    if not thread_ids:
        flash('未選擇任何 Thread', 'warning')
        return redirect(url_for('admin.index', group_id=group_id))

    # Hybrid approach: sync for <=5 threads, async for >5
    if len(thread_ids) <= 5:
        # Synchronous export
        return _generate_sync_pdf_export(project, thread_ids)
    else:
        # Check concurrent task limit before async processing
        from ...models import PDFExportTask
        ongoing_tasks = PDFExportTask.query.filter(
            PDFExportTask.status.in_(['queued', 'running'])
        ).count()
        
        if ongoing_tasks >= 4:  # Max 4 concurrent tasks (2 running + 2 queued)
            flash('⏳ 系統繁忙，目前有多個匯出任務進行中，請稍後再試', 'warning')
            return redirect(url_for('admin.index', group_id=group_id))
        # Asynchronous export - create record FIRST to avoid race condition
        import time
        from ...models import PDFExportTask
        
        # Generate a unique task ID
        import uuid
        task_id = str(uuid.uuid4())
        
        # Create task record with 'queued' status
        task_record = PDFExportTask(
            id=task_id,
            user_id=session['user_id'],
            project_id=project.id,
            thread_count=len(thread_ids),
            status='queued',
            progress_total=len(thread_ids),
            progress_current=0,
            created_at=int(time.time())
        )
        db.session.add(task_record)
        db.session.commit()
        
        # Schedule Huey task with the pre-created task ID
        tasks.generate_batch_pdf_task(project.id, thread_ids, session['user_id'], task_id)
        
        flash(f'已啟動背景任務匯出 {len(thread_ids)} 筆資料', 'info')
        return redirect(url_for('admin.pdf_export_status', task_id=task_id))



def _generate_sync_pdf_export(project, thread_ids):
    """Synchronous PDF generation for small batches"""
    from ...services import pdf_service
    from ... import logic as legacy_services
    from flask import Response
    import io
    import zipfile
    import math
    
    # Helper function
    def get_headers_callback(key):
        return legacy_services.get_headers(key)
        
    from ...utils import generate_pdf_filename, encode_filename_header
    
    CHUNK_SIZE = 50
    
    if len(thread_ids) == 1:
        # Single thread - may return PDF or ZIP depending on size
        thread_id = thread_ids[0]
        thread_data = legacy_services.process_thread(
            {'thread_id': thread_id}, None, None, None, project.api_key, project.id
        )
        
        if not thread_data or not thread_data.get('data'):
            flash('Thread not found', 'error')
            return redirect(url_for('admin.index', group_id=project.id))
        
        messages = thread_data['data']['messages']
        remark = thread_data['data'].get('remark')
        filename = generate_pdf_filename(thread_id, remark)
        
        if len(messages) <= CHUNK_SIZE:
            # Single PDF
            html = render_template('print_view.html', threads=[thread_data['data']])
            html, temp_files = pdf_service.preprocess_html_for_pdf(html, project.id, get_headers_callback)
            try:
                pdf_bytes = pdf_service.generate_pdf_bytes(html)
                return Response(pdf_bytes, mimetype='application/pdf', headers={
                    'Content-Disposition': encode_filename_header(filename)
                })
            finally:
                pdf_service.cleanup_temp_images(temp_files)
        else:
            # Split into ZIP
            chunks = math.ceil(len(messages) / CHUNK_SIZE)
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for i in range(chunks):
                    start = i * CHUNK_SIZE
                    end = start + CHUNK_SIZE
                    chunk_msgs = messages[start:end]
                    
                    chunk_data = thread_data['data'].copy()
                    chunk_data['messages'] = chunk_msgs
                    
                    html = render_template('print_view.html', threads=[chunk_data])
                    html, temp_files = pdf_service.preprocess_html_for_pdf(html, project.id, get_headers_callback)
                    
                    try:
                        pdf_bytes = pdf_service.generate_pdf_bytes(html)
                        # Filename for part
                        base_name = filename.rsplit('.', 1)[0]
                        part_filename = f"{base_name}_part_{i+1}.pdf"
                        zf.writestr(part_filename, pdf_bytes)
                    finally:
                        pdf_service.cleanup_temp_images(temp_files)
            
            zip_buffer.seek(0)
            zip_filename = filename.rsplit('.', 1)[0] + "_split.zip"
            return Response(zip_buffer.getvalue(), mimetype='application/zip', headers={
                'Content-Disposition': encode_filename_header(zip_filename)
            })
    else:
        # Multiple threads - ZIP
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for thread_id in thread_ids:
                try:
                    thread_data = legacy_services.process_thread(
                        {'thread_id': thread_id}, None, None, None, project.api_key, project.id
                    )
                    
                    if not thread_data or not thread_data.get('data'):
                        current_app.logger.warning(f"Thread {thread_id} returned no data: {thread_data}")
                        continue
                    
                    messages = thread_data['data']['messages']
                    remark = thread_data['data'].get('remark')
                    filename = generate_pdf_filename(thread_id, remark)
                    
                    if len(messages) <= CHUNK_SIZE:
                        # Single PDF for this thread
                        html = render_template('print_view.html', threads=[thread_data['data']])
                        html, temp_files = pdf_service.preprocess_html_for_pdf(html, project.id, get_headers_callback)
                        
                        try:
                            pdf_bytes = pdf_service.generate_pdf_bytes(html)
                            zf.writestr(filename, pdf_bytes)
                        finally:
                            pdf_service.cleanup_temp_images(temp_files)
                    else:
                        # Long thread - generate multiple PDFs
                        chunks = math.ceil(len(messages) / CHUNK_SIZE)
                        for i in range(chunks):
                            start = i * CHUNK_SIZE
                            end = start + CHUNK_SIZE
                            chunk_msgs = messages[start:end]
                            
                            chunk_data = thread_data['data'].copy()
                            chunk_data['messages'] = chunk_msgs
                            
                            html = render_template('print_view.html', threads=[chunk_data])
                            html, temp_files = pdf_service.preprocess_html_for_pdf(html, project.id, get_headers_callback)
                            
                            try:
                                pdf_bytes = pdf_service.generate_pdf_bytes(html)
                                base_name = filename.rsplit('.', 1)[0]
                                part_filename = f"{base_name}_part_{i+1}.pdf"
                                zf.writestr(part_filename, pdf_bytes)
                            finally:
                                pdf_service.cleanup_temp_images(temp_files)
                except Exception as e:
                    current_app.logger.error(f"Failed to export thread {thread_id}: {e}", exc_info=True)
                    continue
        
        zip_buffer.seek(0)
        import time
        timestamp = int(time.time())
        return Response(zip_buffer.getvalue(), mimetype='application/zip', headers={
            'Content-Disposition': f'attachment; filename="batch_export_{timestamp}.zip"'
        })


@admin_bp.route('/pdf-export/status/<task_id>')
def pdf_export_status(task_id):
    """Display progress tracking page for async PDF export"""
    if not session.get('user_id'):
        return redirect(url_for('auth.login'))
    
    from ...models import PDFExportTask
    task_record = PDFExportTask.query.get(task_id)
    
    if not task_record:
        flash('任務不存在', 'error')
        return redirect(url_for('admin.index'))
    
    # Permission check: only creator or admin can view
    if session.get('role') != 'admin' and task_record.user_id != session.get('user_id'):
        flash('Permission Denied', 'error')
        return redirect(url_for('admin.index'))
    
    return render_template('admin/pdf_export_status.html', task=task_record)


@admin_bp.route('/api/pdf-export/progress/<task_id>')
def pdf_export_progress(task_id):
    """API endpoint for progress polling"""
    if not session.get('user_id'):
        return {'error': 'Unauthorized'}, 401
    
    from ...models import PDFExportTask
    from flask import jsonify
    
    task_record = PDFExportTask.query.get(task_id)
    
    if not task_record:
        return {'error': 'Task not found'}, 404
    
    # Permission check
    if session.get('role') != 'admin' and task_record.user_id != session.get('user_id'):
        return {'error': 'Permission denied'}, 403
    
    return jsonify({
        'status': task_record.status,
        'progress': {
            'current': task_record.progress_current,
            'total': task_record.progress_total
        },
        'file_path': task_record.file_path if task_record.status == 'completed' else None,
        'error_message': task_record.error_message if task_record.status == 'failed' else None
    })


@admin_bp.route('/pdf-export/download/<task_id>')
def pdf_export_download(task_id):
    """Download completed PDF export file"""
    if not session.get('user_id'):
        return redirect(url_for('auth.login'))
    
    from ...models import PDFExportTask
    from flask import send_file
    import os
    
    task_record = PDFExportTask.query.get(task_id)
    
    if not task_record:
        flash('任務不存在', 'error')
        return redirect(url_for('admin.index'))
    
    # Permission check
    if session.get('role') != 'admin' and task_record.user_id != session.get('user_id'):
        flash('Permission Denied', 'error')
        return redirect(url_for('admin.index'))
    
    if task_record.status != 'completed':
        flash('PDF 尚未生成完成', 'warning')
        return redirect(url_for('admin.pdf_export_status', task_id=task_id))
    
    if not task_record.file_path or not os.path.exists(task_record.file_path):
        flash('檔案不存在或已被清理', 'error')
        return redirect(url_for('admin.index'))
    
    log_audit('Download PDF Export', f"User downloaded PDF export {task_id}")
    
    return send_file(
        task_record.file_path,
        as_attachment=True,
        download_name=f"batch_export_{task_id}.zip"
    )
